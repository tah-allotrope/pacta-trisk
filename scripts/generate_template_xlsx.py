from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_DIR = Path("intake/templates")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

wb = Workbook()

ws_data = wb.active
ws_data.title = "Data"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
example_font = Font(color="666666", italic=True)
error_font = Font(color="FF0000")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

columns = [
    "counterparty_name", "exposure_vnd", "sector_code", "sector_code_system",
    "credit_limit_vnd", "lei", "tax_id", "parent_name", "parent_id", "currency",
]

for col_idx, col_name in enumerate(columns, 1):
    cell = ws_data.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

examples = [
    ["Cong ty CP Nhiet Dien Vinh Tan", 800000000000, "D3511", "VSIC", 960000000000, "", "0301452948", "Tap doan Dien luc Viet Nam (EVN)", "EVN_001", "VND"],
    ["Cong ty CP O to Truong Hai", 450000000000, "2910", "ISIC", 540000000000, "5493000IBP32UQZ0KL24", "0301452949", "Tap doan Truong Hai (THACO)", "THACO_001", "VND"],
    ["Cong ty CP Xi mang VICEM Ha Tien", 320000000000, "2394", "ISIC", 384000000000, "", "0301452950", "Tong cong ty Xi mang VICEM", "VICEM_001", "VND"],
]

for row_idx, example in enumerate(examples, 2):
    for col_idx, value in enumerate(example, 1):
        cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
        cell.font = example_font
        cell.border = thin_border

error_row = 5
invalid_values = ["", -500000000000, "9999", "ISIC", 0, "", "", "", "", "VND"]
ws_data.cell(row=error_row, column=1, value="(Thieu ten ben vay / Missing counterparty name)").font = error_font
ws_data.cell(row=error_row, column=1).border = thin_border
for col_idx, value in enumerate(invalid_values, 1):
    if col_idx > 1:
        cell = ws_data.cell(row=error_row, column=col_idx, value=value)
        cell.font = error_font
        cell.border = thin_border

widths = [42, 18, 14, 20, 18, 28, 14, 42, 14, 12]
for i, w in enumerate(widths, 1):
    ws_data.column_dimensions[get_column_letter(i)].width = w

dv_sector = DataValidation(type="list", formula1='"VSIC,ISIC"', allow_blank=False)
dv_sector.error = "Please select VSIC or ISIC"
dv_sector.errorTitle = "Invalid sector code system"
dv_sector.prompt = "Select the sector code system"
dv_sector.promptTitle = "Sector Code System"
ws_data.add_data_validation(dv_sector)
dv_sector.add("D2:D104")

dv_currency = DataValidation(type="list", formula1='"VND,USD"', allow_blank=True)
dv_currency.error = "Please select VND or USD"
dv_currency.errorTitle = "Invalid currency"
dv_currency.prompt = "Select the currency"
dv_currency.promptTitle = "Currency"
ws_data.add_data_validation(dv_currency)
dv_currency.add("J2:J104")

dv_exposure = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
dv_exposure.error = "Exposure must be >= 0"
dv_exposure.errorTitle = "Invalid exposure"
ws_data.add_data_validation(dv_exposure)
dv_exposure.add("B2:B104")

ws_instructions = wb.create_sheet("Instructions")

instructions = [
    ["BYOL Loanbook Template - Instructions / Hướng dẫn"],
    [""],
    ["ENGLISH"],
    ["This template is for submitting your loanbook data for PACTA + TRISK analysis."],
    ['Fill in the "Data" sheet with your loan portfolio details.'],
    [""],
    ["REQUIRED COLUMNS (must be filled for every loan):"],
    ["  counterparty_name - Legal name of the borrowing entity"],
    ["  exposure_vnd - Outstanding exposure amount in VND (must be >= 0)"],
    ["  sector_code - Industry code (VSIC or ISIC format)"],
    ['  sector_code_system - "VSIC" or "ISIC" (select from dropdown)'],
    ["  credit_limit_vnd - Total credit limit in VND (must be >= 0)"],
    [""],
    ["OPTIONAL COLUMNS:"],
    ["  lei - Legal Entity Identifier (20 characters)"],
    ["  tax_id - Local tax identification number"],
    ["  parent_name - Name of the ultimate parent / group"],
    ["  parent_id - Parent identifier (if known)"],
    ['  currency - "VND" or "USD" (default: VND from dropdown)'],
    [""],
    ["NOTES:"],
    ["  - VSIC codes should include the letter prefix (e.g., D3511)"],
    ["  - ISIC codes should be 4 digits (e.g., 3511)"],
    ["  - All monetary values must be in VND (VND is the default currency)"],
    ["  - Row 5 contains an intentionally invalid example for testing"],
    [""],
    ["---"],
    [""],
    ["TIẾNG VIỆT"],
    ["Mẫu này dùng để gửi dữ liệu danh mục cho vay cho phân tích PACTA + TRISK."],
    ['Điền thông tin vào sheet "Data" với chi tiết danh mục cho vay của ngân hàng.'],
    [""],
    ["CÁC CỘT BẮT BUỘC (phải điền cho mỗi khoản vay):"],
    ["  counterparty_name - Tên pháp nhân của bên vay"],
    ["  exposure_vnd - Dư nợ bằng VND (phải >= 0)"],
    ["  sector_code - Mã ngành (định dạng VSIC hoặc ISIC)"],
    ['  sector_code_system - "VSIC" hoặc "ISIC" (chọn từ dropdown)'],
    ["  credit_limit_vnd - Hạn mức tín dụng bằng VND (phải >= 0)"],
    [""],
    ["CÁC CỘT KHÔNG BẮT BUỘC:"],
    ["  lei - Mã định danh pháp nhân (20 ký tự)"],
    ["  tax_id - Mã số thuế"],
    ["  parent_name - Tên công ty mẹ / tập đoàn"],
    ["  parent_id - Mã định danh công ty mẹ"],
    ['  currency - "VND" hoặc "USD" (mặc định: VND từ dropdown)'],
    [""],
    ["LƯU Ý:"],
    ["  - Mã VSIC bao gồm chữ cái đầu (ví dụ: D3511)"],
    ["  - Mã ISIC gồm 4 chữ số (ví dụ: 3511)"],
    ["  - Tất cả giá trị tiền tệ phải bằng VND (VND là mặc định)"],
    ["  - Hàng thứ 5 chứa dữ liệu lỗi cố ý để kiểm tra"],
]

for row_idx, line in enumerate(instructions, 1):
    cell = ws_instructions.cell(row=row_idx, column=1, value=line[0] if line else "")
    if row_idx == 1:
        cell.font = Font(bold=True, size=14, color="1F4E79")

ws_instructions.column_dimensions["A"].width = 80

output_path = OUTPUT_DIR / "loanbook_template.xlsx"
wb.save(str(output_path))
print(f"XLSX template created at {output_path}")
